#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bid materials — 标书素材库管家（吸收 elevator-docflow-skill 思路，按标书通用场景改造）

功能:
  init    初始化素材库: 8大分类目录 + keywords.json 规则库 + 必备材料清单 + 素材清单xlsx
  analyze 扫描素材库 → 规则匹配 → 生成 plan.md 预览（低置信度标 needs_review）
  apply   执行 plan: 移动+重命名（自动备份到 .backup/）+ 写回清单xlsx
  status  对照投标必备材料清单, 标出 已有/缺失
  learn   处理"待整理"文件: AI建议(可选) → 人确认 → 写回 keywords.json, 规则越用越长

设计哲学（源自《垂直于电梯公司的AI工具》一文, 改造为通用标书场景）:
  - 纯规则模式 0 token: 文件名关键词匹配覆盖 90% 常见文件
  - 规则是活的: learn 确认后写回 keywords.json, 换机器拷贝即继承
  - 每一步可审计: plan.md 记录每个文件的源路径/目标/规则/置信度
  - 安全第一: apply 前自动备份, 删除只进回收站
"""

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime

VERSION = "1.0.0"

# ─────────────────────────── 默认配置 ───────────────────────────

DEFAULT_CATEGORIES = [
    "01-企业证照",  # 营业执照/ISO/许可证/信用报告/能效证书
    "02-财务资料",  # 审计报告/财报/完税证明/银行资信
    "03-合同业绩",  # 中标通知书/合同/验收报告/业主证明
    "04-人员资质",  # 职称证书/执业资格/社保/身份证/简历
    "05-技术资料",  # 专利/检测报告/型式试验/技术方案/工艺
    "06-方案售后",  # 安装计划/售后承诺/维保方案/应急预案
    "07-品牌模板",  # logo/商标/授权书/委托书/盖章模板
    "08-荣誉奖项",  # 获奖证书/表扬信/锦旗
    "00-待整理",    # needs_review 兜底
]

# 内置默认规则库: 关键词 → 分类 + 命名模板
# name_template 中的 {公司}/{项目}/{姓名}/{证书} 由 learn/AI 模式提取填充, 缺省留空
DEFAULT_KEYWORDS = {
    "01-企业证照": [
        {"keys": ["营业执照", "统一社会信用代码"], "name": "营业执照（{公司}）"},
        {"keys": ["iso9001", "iso 9001", "质量管理体系"], "name": "ISO9001质量管理体系认证"},
        {"keys": ["iso14001", "iso 14001", "环境管理体系"], "name": "ISO14001环境管理体系认证"},
        {"keys": ["iso45001", "iso 45001", "职业健康安全"], "name": "ISO45001职业健康安全认证"},
        {"keys": ["iso27001", "信息安全管理"], "name": "ISO27001信息安全管理认证"},
        {"keys": ["信用中国", "信用报告", "信用等级"], "name": "企业信用报告"},
        {"keys": ["特种设备", "制造许可证", "生产许可证"], "name": "特种设备制造许可证"},
        {"keys": ["高新技术企业", "高企"], "name": "高新技术企业证书"},
        {"keys": ["安全生产许可证"], "name": "安全生产许可证"},
        {"keys": ["建筑业企业资质", "施工资质"], "name": "建筑业企业资质证书"},
        {"keys": ["能效", "节能认证"], "name": "能效证书"},
        {"keys": ["食品经营许可证", "卫生许可证"], "name": "经营许可证"},
        {"keys": ["软件著作权", "软著"], "name": "软件著作权证书"},
        {"keys": ["商标注册证"], "name": "商标注册证"},
        {"keys": ["开户许可", "基本存款账户"], "name": "开户许可证"},
    ],
    "02-财务资料": [
        {"keys": ["审计报告"], "name": "审计报告（{年份}）"},
        {"keys": ["财务报表", "资产负债表"], "name": "财务报表（{年份}）"},
        {"keys": ["利润表", "损益表"], "name": "利润表（{年份}）"},
        {"keys": ["现金流量表"], "name": "现金流量表（{年份}）"},
        {"keys": ["完税证明", "纳税证明", "税收"], "name": "完税证明（{年份}）"},
        {"keys": ["银行资信", "资信证明"], "name": "银行资信证明"},
        {"keys": ["验资报告"], "name": "验资报告"},
    ],
    "03-合同业绩": [
        {"keys": ["中标通知书", "中标公告"], "name": "中标通知书（{项目}）"},
        {"keys": ["合同"], "name": "合同证明（{项目}）"},
        {"keys": ["验收报告", "验收单", "竣工验收"], "name": "验收报告（{项目}）"},
        {"keys": ["业绩证明", "业主评价", "用户评价", "履约证明"], "name": "业绩证明（{项目}）"},
        {"keys": ["结算单", "结算证明"], "name": "结算证明（{项目}）"},
    ],
    "04-人员资质": [
        {"keys": ["高级工程师", "高工"], "name": "高级工程师-{姓名}"},
        {"keys": ["中级工程师", "工程师证书"], "name": "工程师-{姓名}"},
        {"keys": ["注册建造师", "一级建造师", "二级建造师"], "name": "注册建造师-{姓名}"},
        {"keys": ["职称", "资格证"], "name": "职称证书-{姓名}"},
        {"keys": ["特种作业", "安全员", "安全生产考核"], "name": "安全证书-{姓名}"},
        {"keys": ["焊工", "电工证", "叉车", "起重", "特种设备作业"], "name": "特种作业证-{姓名}"},
        {"keys": ["驾驶证", "驾照"], "name": "驾驶证-{姓名}"},
        {"keys": ["身份证"], "name": "身份证-{姓名}"},
        {"keys": ["社保证明", "社保"], "name": "社保证明-{姓名}"},
        {"keys": ["毕业证", "学历"], "name": "学历证书-{姓名}"},
        {"keys": ["简历"], "name": "简历-{姓名}"},
    ],
    "05-技术资料": [
        {"keys": ["专利证书", "专利"], "name": "专利证书-{项目}"},
        {"keys": ["检测报告", "检验报告", "型式试验"], "name": "检测报告（{项目}）"},
        {"keys": ["技术方案", "实施方案"], "name": "技术方案（{项目}）"},
        {"keys": ["产品说明书", "使用手册"], "name": "产品说明书-{项目}"},
        {"keys": ["工艺", "作业指导书"], "name": "工艺文件-{项目}"},
        {"keys": ["测试报告", "试验报告"], "name": "测试报告（{项目}）"},
    ],
    "06-方案售后": [
        {"keys": ["安装计划", "施工组织"], "name": "安装计划（{项目}）"},
        {"keys": ["售后服务", "售后承诺", "质保"], "name": "售后承诺（{项目}）"},
        {"keys": ["维保", "维护保养", "保修"], "name": "维保方案（{项目}）"},
        {"keys": ["应急预案", "安全方案"], "name": "应急预案（{项目}）"},
        {"keys": ["培训方案", "培训计划"], "name": "培训方案（{项目}）"},
    ],
    "07-品牌模板": [
        {"keys": ["logo", "标志", "标识"], "name": "公司Logo"},
        {"keys": ["授权书", "委托书"], "name": "授权书（{项目}）"},
        {"keys": ["模板", "空白表"], "name": "模板-{名称}"},
        {"keys": ["公章", "印模", "骑缝章"], "name": "公章印模"},
    ],
    "08-荣誉奖项": [
        {"keys": ["获奖", "荣誉", "优秀", "先进", "锦旗", "表扬信"], "name": "荣誉证书-{名称}"},
        {"keys": ["3a", "aaa信用"], "name": "AAA信用等级证书"},
    ],
}

# 投标必备材料清单（status 对照基准, 可按公司情况编辑）
DEFAULT_CHECKLIST = [
    "营业执照",
    "ISO9001质量管理体系认证",
    "审计报告（近三年）",
    "完税证明",
    "企业信用报告",
    "安全生产许可证",
    "特种设备制造许可证",
    "主要人员资质证书",
    "合同业绩证明",
    "中标通知书",
    "银行资信证明",
    "荣誉证书",
]

# 不需要整理的隐藏/系统目录
SKIP_DIRS = {".backup", ".recycle", ".git", "__pycache__"}

# 元文件（工具自己生成的, 不参与扫描/整理）
META_FILES = {"keywords.json", "checklist.json", "素材清单.xlsx", "plan.md", "plan.json"}


def is_meta_file(fn):
    """判断是否工具元文件（json规则/bak/tmp/清单/计划）"""
    if fn in META_FILES:
        return True
    lower = fn.lower()
    return lower.endswith(".bak") or lower.endswith(".tmp") or fn.startswith(".")


# ─────────────────────────── 工具函数 ───────────────────────────

def ts():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ts_dir():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def log(msg):
    print(msg)


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)
    return path


def load_json(path, default):
    if os.path.isfile(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log(f"⚠️  规则文件解析失败 {path}: {e}，使用默认规则")
    return default


def save_json(path, data):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    # 写前备份旧文件（竞品思路: 更新规则前自动备份）
    if os.path.isfile(path):
        shutil.copy2(path, path + ".bak")
    shutil.move(tmp, path)


def ext_of(filename):
    return os.path.splitext(filename)[1].lower()


def strip_ext(filename):
    return os.path.splitext(filename)[0]


def pdf_first_page_text(path, max_chars=300):
    """尝试提取 PDF 第一页文本用于内容匹配（可选，失败静默）"""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(path)
        if doc.page_count > 0:
            text = doc[0].get_text()[:max_chars]
            doc.close()
            return text
    except Exception:
        pass
    return ""


# ─────────────────────────── init ───────────────────────────

def cmd_init(args):
    base = os.path.abspath(args.dir)
    ensure_dir(base)

    # 1. 创建分类目录
    created = []
    for cat in DEFAULT_CATEGORIES:
        d = os.path.join(base, cat)
        if not os.path.isdir(d):
            ensure_dir(d)
            created.append(cat)

    # 2. 规则库 keywords.json
    kf = os.path.join(base, "keywords.json")
    if not os.path.isfile(kf):
        save_json(kf, DEFAULT_KEYWORDS)
        log(f"📄 已生成规则库: {kf}")

    # 3. 必备材料清单 checklist.json
    cf = os.path.join(base, "checklist.json")
    if not os.path.isfile(cf):
        save_json(cf, DEFAULT_CHECKLIST)
        log(f"📄 已生成必备材料清单: {cf}")

    # 4. 素材清单 xlsx（空模板）
    xf = os.path.join(base, "素材清单.xlsx")
    if not os.path.isfile(xf):
        _write_manifest(base, [])
        log(f"📄 已生成素材清单: {xf}")

    # 5. 回收站/备份目录
    ensure_dir(os.path.join(base, ".backup"))
    ensure_dir(os.path.join(base, ".recycle"))

    log(f"\n✅ 素材库初始化完成: {base}")
    log(f"   分类目录 {len(DEFAULT_CATEGORIES)} 个: {' / '.join(DEFAULT_CATEGORIES)}")
    log(f"   下一步: 把素材文件丢进 {base} 根目录（或任意子目录），然后跑:")
    log(f"     bid materials analyze {base}")


# ─────────────────────────── 规则匹配 ───────────────────────────

def match_rule(filename, keywords):
    """按关键词匹配文件 → (分类, 命名模板, 置信度)。文件名+PDF首页文本双重匹配。"""
    name_lower = filename.lower()
    best = None
    best_score = 0
    for cat, rules in keywords.items():
        for rule in rules:
            for kw in rule["keys"]:
                if kw.lower() in name_lower:
                    score = len(kw)
                    if score > best_score:
                        best = (cat, rule["name"], score)
                        best_score = score
    if best:
        # 命中即 0.72 起, 关键词越长置信度越高（避免正确匹配被误标 needs_review）
        conf = min(0.98, 0.72 + best_score / 50.0)
        return best[0], best[1], conf
    return None, None, 0.0


def extract_entities(filename, category, name_template):
    """从文件名提取 {公司}/{项目}/{姓名}/{年份}/{名称} 等实体（启发式, 供命名用）"""
    name = strip_ext(filename)
    entities = {}

    # 年份: 20XX
    m = re.search(r"(20[12]\d)", name)
    if m:
        entities["年份"] = m.group(1)

    # 姓名（仅证书/简历/身份证类模板需要）——按优先级尝试
    if "姓名" in name_template:
        # 模式1: 分隔符后的人名: "身份证-张三" / "高级工程师-李四" / "简历-王五"
        m = re.search(r"[\-—_（(]([\u4e00-\u9fa5]{2,3})(?:证书|复印件)?$", name)
        if not m:
            # 模式2: 证件词紧邻前缀人名: "张三身份证" / "李四简历"（排除非人名常见词）
            m = re.search(r"(?<![\u4e00-\u9fa5])(?!员工|个人|公司|求职|本人|工作)([\u4e00-\u9fa5]{2,3})(?:身份证|简历|毕业证|社保证明|驾驶证)", name)
        if not m:
            # 模式3: 职称后紧邻: "高级工程师李四"
            m = re.search(r"(?:工程师|职称|建造师|技师)([\u4e00-\u9fa5]{2,3})$", name)
        if m:
            cand = m.group(1)
            # 排除文件类型词
            if cand not in ("高级工程师", "中级工程师", "注册建造师", "职称证书", "证书", "身份证", "社保证明"):
                entities["姓名"] = cand

    # 公司名: "XX有限公司"
    m = re.search(r"([\u4e00-\u9fa5]{2,20}有限公司)", name)
    if m:
        entities["公司"] = m.group(1)

    # 项目/产品名
    if "项目" in name_template or "产品" in name_template or "名称" in name_template:
        # 模式1: 分隔符后整词: "中标通知书-某项目" / "专利证书-智能安防系统"
        m = re.search(r"[\-—_（(]([\u4e00-\u9fa5A-Za-z0-9]{2,20})$", name)
        if not m:
            # 模式2: XX项目/工程/采购/系统
            m = re.search(r"([\u4e00-\u9fa5A-Za-z0-9]{2,20}?(?:项目|工程|采购|系统))", name)
        if m:
            entities["项目"] = m.group(1)

    # 荣誉/奖项名称: "2024年度优秀供应商获奖证书" → 提取"获奖/荣誉"前部分
    if "名称" in name_template and "项目" not in entities:
        m = re.search(r"^(.{2,20}?)(?:获奖|荣誉)", name)
        if m:
            entities["名称"] = m.group(1)

    return entities


def render_name(template, entities):
    """按模板渲染文件名（缺实体则省略对应占位）"""
    result = template
    for key, val in entities.items():
        result = result.replace("{" + key + "}", val)
    # 清除未填充占位符: {公司}/{项目}/{姓名} 等
    result = re.sub(r"\{[^}]+\}", "", result)
    # 清理双括号/多余符号
    result = re.sub(r"（（", "（", result).replace("（）", "")
    result = re.sub(r"[\-_—]{2,}", "-", result).strip("-—_ ")
    return result


# ─────────────────────────── analyze ───────────────────────────

def cmd_analyze(args):
    base = os.path.abspath(args.dir)
    if not os.path.isdir(base):
        log(f"❌ 素材库不存在: {base}，先跑 bid materials init")
        sys.exit(1)

    keywords = load_json(os.path.join(base, "keywords.json"), DEFAULT_KEYWORDS)
    checklist = load_json(os.path.join(base, "checklist.json"), DEFAULT_CHECKLIST)

    # 收集所有待整理文件（跳过分类目录/隐藏目录/备份）
    files = []
    for root, dirs, fnames in os.walk(base):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS and not d.startswith(".")]
        for fn in sorted(fnames):
            if is_meta_file(fn):
                continue
            full = os.path.join(root, fn)
            rel = os.path.relpath(full, base)
            files.append((rel, full, fn))

    if not files:
        log(f"📭 素材库为空（或没有待整理文件）: {base}")
        return

    # 规则匹配
    plan = []
    unmatched = 0
    for rel, full, fn in files:
        cat, tmpl, conf = match_rule(fn, keywords)
        # PDF 内容兜底匹配（文件名没命中时）
        if not cat and fn.lower().endswith(".pdf"):
            txt = pdf_first_page_text(full)
            if txt:
                cat, tmpl, conf = match_rule(fn + " " + txt, keywords)
        if not cat:
            cat, tmpl, conf = "00-待整理", None, 0.0
            unmatched += 1
        entities = extract_entities(fn, cat, tmpl or "")
        new_name = render_name(tmpl, entities) + ext_of(fn) if tmpl else fn
        needs_review = conf < args.min_conf or cat == "00-待整理"
        plan.append({
            "src": rel, "src_full": full, "filename": fn,
            "category": cat, "template": tmpl, "confidence": round(conf, 2),
            "new_name": new_name, "needs_review": needs_review,
        })

    # 预览
    log(f"🔍 扫描完成: {len(files)} 个文件, 规则命中 {len(files) - unmatched}, 待人工确认 {unmatched}")
    log(f"   置信度阈值: {args.min_conf}\n")
    log(f"{'状态':<6}{'分类':<14}{'置信度':<8}文件")
    log("-" * 80)
    for p in sorted(plan, key=lambda x: (x["needs_review"], -x["confidence"])):
        flag = "⚠️待审" if p["needs_review"] else "✅"
        log(f"{flag:<6}{p['category']:<14}{p['confidence']:<8}{p['filename']}")

    # 写 plan.md（审计日志）
    plan_path = os.path.join(base, "plan.md")
    lines = [f"# 素材整理计划（{ts()}）\n",
             f"- 扫描文件数: {len(files)}",
             f"- 规则命中: {len(files) - unmatched}",
             f"- 待人工确认: {unmatched}",
             f"- 阈值: {args.min_conf}\n"]
    for p in sorted(plan, key=lambda x: (x["needs_review"], -x["confidence"])):
        mark = "⚠️ needs_review" if p["needs_review"] else "✅"
        lines.append(f"- [{mark}] {p['filename']}")
        lines.append(f"    源路径: {p['src']}")
        lines.append(f"    分类: {p['category']} | 命名: {p['new_name']} | 置信度: {p['confidence']}")
        if p["template"]:
            lines.append(f"    规则: {p['template']}")
    with open(plan_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    log(f"\n📁 计划已写入: {plan_path}（完整审计: 源路径/目标/规则/置信度）")

    # 保存 plan.json 供 apply 使用
    with open(os.path.join(base, "plan.json"), "w", encoding="utf-8") as f:
        json.dump(plan, f, ensure_ascii=False, indent=2)

    n_review = len([p for p in plan if p["needs_review"]])
    if n_review:
        log(f"\n⚠️  {n_review} 个文件需要人工确认（低置信度/未匹配），确认后跑:")
        log(f"   bid materials apply {base}")
        log(f"   或先用 bid materials learn {base} 教它认识新文件")
    else:
        log(f"\n✅ 全部文件已匹配，直接执行:")
        log(f"   bid materials apply {base}")


# ─────────────────────────── apply ───────────────────────────

def _write_manifest(base, rows):
    """写素材清单 xlsx（openpyxl）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "素材清单"
    headers = ["序号", "文件名", "原路径", "分类", "状态", "备注", "更新时间"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)

    widths = [6, 40, 30, 14, 10, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(os.path.join(base, "素材清单.xlsx"))


def cmd_apply(args):
    base = os.path.abspath(args.dir)
    plan_path = os.path.join(base, "plan.json")
    if not os.path.isfile(plan_path):
        log(f"❌ 没有 plan.json，先跑: bid materials analyze {base}")
        sys.exit(1)

    with open(plan_path, "r", encoding="utf-8") as f:
        plan = json.load(f)

    # learn 可能已处理部分文件（源文件已移动走），先过滤掉
    before = len(plan)
    plan = [p for p in plan if os.path.isfile(os.path.join(base, p["src"]))]
    if len(plan) < before:
        log(f"ℹ️  跳过 {before - len(plan)} 个已被 learn/手动处理过的文件")

    review_items = [p for p in plan if p["needs_review"]]
    if review_items and not args.force:
        log(f"⚠️  {len(review_items)} 个文件标了 needs_review（低置信度/未匹配）。")
        log(f"   用 --force 强制执行，或先跑 bid materials learn {base} 处理")
        for p in review_items[:10]:
            log(f"     - {p['filename']} → {p['category']}")
        sys.exit(1)

    # 备份目录
    bak_dir = ensure_dir(os.path.join(base, ".backup", ts_dir()))

    moved, skipped = 0, 0
    rows = []
    for p in plan:
        src_full = os.path.join(base, p["src"])
        if not os.path.isfile(src_full):
            skipped += 1
            continue
        cat_dir = ensure_dir(os.path.join(base, p["category"]))
        dst = os.path.join(cat_dir, p["new_name"])
        # 重名处理
        if os.path.exists(dst):
            stem, e = os.path.splitext(p["new_name"])
            dst = os.path.join(cat_dir, f"{stem}_1{e}")
        # 备份原文件
        try:
            shutil.copy2(src_full, os.path.join(bak_dir, os.path.basename(src_full)))
        except Exception:
            pass
        # 移动
        shutil.move(src_full, dst)
        status = "待整理" if p["category"] == "00-待整理" else "已有"
        rows.append([
            len(rows) + 1, os.path.basename(dst), p["src"], p["category"],
            status, "自动整理" if status == "已有" else "需人工归类", ts(),
        ])
        moved += 1
        if args.verbose:
            log(f"  📦 {os.path.basename(src_full)} → {p['category']}/{os.path.basename(dst)}")

    # 写回素材清单
    _write_manifest(base, rows)
    log(f"✅ 执行完成: 移动 {moved} 个, 跳过 {skipped} 个")
    log(f"   备份目录: {bak_dir}")
    log(f"   素材清单已更新: {os.path.join(base, '素材清单.xlsx')}")
    log(f"   查看缺什么: bid materials status {base}")


# ─────────────────────────── status ───────────────────────────

def cmd_status(args):
    base = os.path.abspath(args.dir)
    if not os.path.isdir(base):
        log(f"❌ 素材库不存在: {base}")
        sys.exit(1)

    checklist = load_json(os.path.join(base, "checklist.json"), DEFAULT_CHECKLIST)
    manifest = os.path.join(base, "素材清单.xlsx")
    have_names = []
    if os.path.isfile(manifest):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(manifest)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[4] == "已有":
                    have_names.append(str(row[1]))
            wb.close()
        except Exception as e:
            log(f"⚠️  清单读取失败: {e}")

    # 也扫描目录兜底
    have_set = set()
    for cat in os.listdir(base):
        cat_dir = os.path.join(base, cat)
        if os.path.isdir(cat_dir) and cat not in SKIP_DIRS:
            for fn in os.listdir(cat_dir):
                have_set.add(strip_ext(fn))
    have_set.update(have_names)

    log(f"📋 投标必备材料状态（{len(checklist)} 项）\n")
    ok, missing = 0, 0
    for item in checklist:
        # 模糊匹配: 清单项关键词是否出现在已有文件里
        found = any(kw.lower() in h.lower() for h in have_set for kw in item.split("（")[0].split("/"))
        if found:
            log(f"  ✅ {item}")
            ok += 1
        else:
            log(f"  ❌ {item}")
            missing += 1

    log(f"\n  已有 {ok}/{len(checklist)}，缺失 {missing} 项")
    if missing:
        log(f"  ⚠️  开标前优先补齐缺失材料！")

    # 目录分布
    log(f"\n📁 素材库分布:")
    for cat in sorted(DEFAULT_CATEGORIES):
        d = os.path.join(base, cat)
        if os.path.isdir(d):
            n = len([f for f in os.listdir(d) if not f.startswith(".")])
            log(f"  {cat}: {n} 个文件")


# ─────────────────────────── learn ───────────────────────────

def llm_suggest(filename, llm_client=None):
    """AI 模式: 建议分类与命名（可选, 需要配置大模型API）"""
    if llm_client is None:
        return None
    prompt = (
        f"你是投标文件管理专家。文件《{filename}》需要归档到标书素材库。\n"
        f"可选分类: {', '.join(DEFAULT_CATEGORIES)}\n"
        f"请返回 JSON: {{\"category\": \"分类名\", \"new_name\": \"建议文件名(不带扩展名)\", \"keywords\": [\"2-3个识别关键词\"]}}\n"
        f"只返回 JSON。"
    )
    try:
        resp = llm_client.chat(prompt)
        # 提取 JSON
        m = re.search(r"\{.*\}", resp, re.S)
        if m:
            data = json.loads(m.group(0))
            return data
    except Exception as e:
        log(f"  ⚠️  AI建议失败: {e}")
    return None


def collect_pending(base, keywords):
    """收集待学习文件：根目录散落文件 + 00-待整理目录（只收未匹配的）"""
    pending = []
    # 根目录散落文件
    for fn in sorted(os.listdir(base)):
        full = os.path.join(base, fn)
        if not os.path.isfile(full) or is_meta_file(fn):
            continue
        cat, tmpl, conf = match_rule(fn, keywords)
        if not cat:  # 未匹配任何规则 → 待学习
            pending.append((full, fn))
    # 00-待整理 目录
    pend_dir = os.path.join(base, "00-待整理")
    if os.path.isdir(pend_dir):
        for fn in sorted(os.listdir(pend_dir)):
            full = os.path.join(pend_dir, fn)
            if os.path.isfile(full) and not is_meta_file(fn):
                pending.append((full, fn))
    return pending


def cmd_learn(args):
    base = os.path.abspath(args.dir)
    keywords = load_json(os.path.join(base, "keywords.json"), DEFAULT_KEYWORDS)
    kf = os.path.join(base, "keywords.json")

    pending = collect_pending(base, keywords)

    if not pending:
        log("🎉 没有待整理文件")
        return

    # 加载 LLM（可选）
    llm = None
    if args.llm:
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from llm_client import LLMClient
            llm = LLMClient()
            log("🤖 AI 模式开启（建议由模型给出，你确认）")
        except Exception as e:
            log(f"⚠️  LLM 加载失败({e})，退回纯人工模式")

    log(f"🧠 自适应学习: 共 {len(pending)} 个待整理文件\n")
    for full, fn in pending:
        log(f"  ▶ {fn}")
        suggestion = None
        if llm:
            suggestion = llm_suggest(fn, llm)
        if suggestion:
            log(f"    🤖 AI建议: 分类[{suggestion.get('category')}] 命名[{suggestion.get('new_name')}]")
        log(f"    当前未匹配任何规则")
        ans = input("    确认归档? [回车跳过 / 输入分类编号或名称 / d删除 / q退出]: ").strip()
        if ans.lower() in ("q", "quit"):
            break
        if ans.lower() in ("d", "del", "删除"):
            os.rename(full, os.path.join(base, ".recycle", fn))
            log(f"    🗑️ 已移入回收站: .recycle/{fn}")
            continue
        if not ans:
            continue
        # 用户输入分类（编号或名称）
        cat = ans
        if ans.isdigit() and 1 <= int(ans) <= len(DEFAULT_CATEGORIES):
            cat = DEFAULT_CATEGORIES[int(ans) - 1]
        # 建议新关键词（从文件名提取核心词）
        base_name = strip_ext(fn)
        kw = input(f"    识别关键词（回车用文件名核心词「{base_name}」）: ").strip() or base_name
        new_name = input(f"    目标文件名（回车保持原名）: ").strip() or base_name

        # 写回规则库（规则生长）
        rules = keywords.setdefault(cat, [])
        # 查重
        if not any(kw in rule["keys"] for rule in rules):
            rules.append({
                "keys": [kw],
                "name": new_name + "（{项目}）" if "项目" in new_name else new_name,
            })
            save_json(kf, keywords)
            log(f"    📝 新规则已写入 keywords.json: [{kw}] → {cat}")
            log(f"       下次 analyze 自动识别同类文件，规则库现在 {len(keywords)} 个分类")
            # 立即移动
            cat_dir = ensure_dir(os.path.join(base, cat))
            dst = os.path.join(cat_dir, os.path.basename(fn))
            shutil.move(full, dst)
            log(f"    📦 已移动: {fn} → {cat}/")
        else:
            log(f"    ⏭️  规则已存在，跳过")
        print()

    log("✅ 学习完成。规则是活的——每确认一次，它就更懂你的文件。")


# ─────────────────────────── main ───────────────────────────

def main():
    parser = argparse.ArgumentParser(
        prog="bid materials",
        description="📦 标书素材库管家 — 8大分类自动整理 + 规则自适应学习 + 必备材料清单",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", help="子命令")

    p = sub.add_parser("init", help="初始化素材库（分类目录+规则库+清单）")
    p.add_argument("dir", nargs="?", default=".", help="素材库目录 (默认: 当前目录)")

    p = sub.add_parser("analyze", help="扫描并生成整理计划 plan.md")
    p.add_argument("dir", nargs="?", default=".", help="素材库目录")
    p.add_argument("--min-conf", type=float, default=0.7, help="置信度阈值，低于此值标 needs_review (默认0.7)")

    p = sub.add_parser("apply", help="执行整理计划（移动+重命名+写回清单）")
    p.add_argument("dir", nargs="?", default=".", help="素材库目录")
    p.add_argument("--force", action="store_true", help="强制执行（跳过 needs_review 确认）")
    p.add_argument("-v", "--verbose", action="store_true", help="显示每个文件的移动明细")

    p = sub.add_parser("status", help="对照必备材料清单，标出已有/缺失")
    p.add_argument("dir", nargs="?", default=".", help="素材库目录")

    p = sub.add_parser("learn", help="自适应学习：教它认识新文件（规则写回 keywords.json）")
    p.add_argument("dir", nargs="?", default=".", help="素材库目录")
    p.add_argument("--llm", action="store_true", help="开启 AI 建议（需要配置大模型API）")

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        sys.exit(0)

    cmds = {"init": cmd_init, "analyze": cmd_analyze, "apply": cmd_apply,
            "status": cmd_status, "learn": cmd_learn}
    cmds[args.command](args)


if __name__ == "__main__":
    main()
